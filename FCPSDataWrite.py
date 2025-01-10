from difflib import get_close_matches  # 从difflib模块导入get_close_matches函数，用于查找最接近的匹配项
import re  # 导入正则表达式模块
import os  # 导入操作系统模块
import pandas as pd  # 导入Pandas模块，用于数据处理
import openpyxl  # 导入Openpyxl模块，用于处理Excel文件
from docx import Document  # 从docx模块导入Document类，用于处理Word文件
from datetime import datetime  # 从datetime模块导入datetime类，用于处理日期时间
from decimal import Decimal  # 从decimal模块导入Decimal类，用于精确浮点数运算

# 定义函数，查找符合日期模式的文件
def find_date_files(folder_path, pattern_str):
    pattern = re.compile(pattern_str)  # 编译正则表达式模式
    return [os.path.join(root, filename) for root, _, files in os.walk(folder_path) for filename in files if pattern.match(filename)]  # 返回符合模式的文件路径列表

# 定义函数，将浮点数格式化为字符串
def format_float_as_string(float_value, decimal_places=2):
    float_str = f"{float_value:.{decimal_places}f}".rstrip('0').rstrip('.')  # 格式化浮点数并去除多余的零
    return float_str  # 返回格式化后的字符串

# 定义函数，判断数组是否按升序排列
def ascending_order_judgment(arr):
    if arr.size == 0:  # 如果数组为空，返回0
        return 0
    for i in range(1, len(arr)):  # 遍历数组
        if arr[i] != arr[i - 1] + 1:  # 如果当前元素不等于前一个元素加1
            return [arr[i - 1], i]  # 返回前一个元素和当前索引
    return None  # 如果数组按升序排列，返回None

# 定义Word文档路径和文件名模式
word_doc_path = "/Users/hang/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/237e6ba0b2b7af52a974319f16a23be7/Message/MessageTemp/c1c30910da9401121316d3c0d92ec2ed/File/"
word_str = r'^\d{4}-\d{2}-\d{2}RB\.docx$'
word_docx_path = find_date_files(word_doc_path, word_str)  # 查找符合模式的Word文件路径
doc = Document(word_docx_path[0])  # 打开第一个符合条件的Word文件

word_content = '\n'.join(para.text for para in doc.paragraphs)  # 获取Word文档的内容

# 定义正则表达式模式，用于匹配日期、总服务费和储能电站服务费
pattern_datetime = re.compile(r"\d{4}年\d{2}月\d{2}日")
pattern_total_amount = re.compile(r"\d{4}年\d{2}月\d{2}日调峰辅助服务市场产生总服务费(\d+(\.\d+)?)万元。")
pattern_energy_storage = re.compile(r"，储能电站总服务费为(\d+(\.\d+)?)万元（占比(\d+(\.\d+)?)%）。")

empty_2d_array = [[None] * 15 for _ in range(1)]  # 创建一个空的二维数组

date_match = pattern_datetime.search(word_content)  # 在Word内容中查找日期
if date_match:
    date_obj = datetime.strptime(date_match.group(), "%Y年%m月%d日")  # 将匹配到的日期字符串转换为日期对象
    empty_2d_array[0][0] = date_obj.strftime("%Y/%m/%d").replace("/0", "/")  # 将日期格式化为字符串并存入数组

date_str_processed = date_match.group()[0:4] + "/" + date_match.group()[5:7] + "/" + date_match.group()[8:10]  # 处理日期字符串
date_obj = datetime.strptime(date_str_processed, "%Y/%m/%d")  # 将处理后的日期字符串转换为日期对象
weekday_name = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"][date_obj.weekday()]  # 获取星期名称
empty_2d_array[0][1] = weekday_name  # 将星期名称存入数组

total_amount_match = pattern_total_amount.search(word_content)  # 在Word内容中查找总服务费
if total_amount_match:
    empty_2d_array[0][5] = float(total_amount_match.group(1))  # 将匹配到的总服务费存入数组

energy_storage_match = pattern_energy_storage.search(word_content)  # 在Word内容中查找储能电站服务费
if energy_storage_match:
    empty_2d_array[0][6] = float(Decimal(energy_storage_match.group(3)) / Decimal(100))  # 将匹配到的储能电站服务费占比存入数组

# 定义Excel文档路径和文件名模式
excel_doc_path = "/Users/hang/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/237e6ba0b2b7af52a974319f16a23be7/Message/MessageTemp/c1c30910da9401121316d3c0d92ec2ed/File/"
excel_report_str = r'福冲储能电站运行日报.*.xlsx'
excel_settlement_str = r'.*深度调峰周期结算数据.xlsx'
matched_report_files = find_date_files(excel_doc_path, excel_report_str)  # 查找符合模式的运行日报文件路径
matched_settlement_str = find_date_files(excel_doc_path, excel_settlement_str)  # 查找符合模式的结算数据文件路径

formatted_date_simple = date_match.group()[date_match.group().find("年") + 1:]
sheet_name = formatted_date_simple  # 将处理后的日期字符串作为工作表名称

if not matched_report_files:
    raise FileNotFoundError("No matching report files found.")
workbook = openpyxl.load_workbook(matched_report_files[0])  # 打开第一个符合条件的运行日报文件
closest_matches = get_close_matches(sheet_name, workbook.sheetnames, n=1, cutoff=0.6)
if not closest_matches:
    raise ValueError(f"No close match found for sheet '{sheet_name}' in the workbook.")  # 如果没有找到接近的工作表，抛出异常
sheet = workbook[closest_matches[0]]  # 获取最接近的工作表

empty_2d_array[0][2] = sheet['F4'].value  # 获取单元格F4的值并存入数组
empty_2d_array[0][3] = sheet['H4'].value.replace("℃", "")  # 获取单元格H4的值并去掉所有的℃后存入数组
empty_2d_array[0][4] = sheet['J4'].value  # 获取单元格J4的值并存入数组

empty_2d_array[0][7] = float(Decimal(sheet['A6'].value))  # 获取单元格A6的值并转换为浮点数后存入数组
empty_2d_array[0][8] = float(Decimal(sheet['B6'].value))  # 获取单元格B6的值并转换为浮点数后存入数组

df = pd.read_excel(matched_settlement_str[0], header=2)  # 读取第一个符合条件的结算数据文件
non_zero_I = df['计算深调电量(MWH)'][df['计算深调电量(MWH)'] != 0]  # 获取非零的深调电量
if non_zero_I.size == 0:
    empty_2d_array[0][9] = 0  # 如果深调电量为空，存入0
else:
    empty_2d_array[0][9] = float(Decimal(non_zero_I.sum()) * Decimal(1000))  # 计算总深调电量并存入数组

non_zero_D = df['交易时段'][df['计算深调电量(MWH)'] != 0]  # 获取非零的交易时段
D_values = ascending_order_judgment(non_zero_D.values)  # 判断交易时段是否按升序排列
if D_values is None:
    empty_2d_array[0][10] = f"{non_zero_D.values[0]}-{non_zero_D.values[-1]}"  # 如果按升序排列，存入数组
elif D_values == 0:
    empty_2d_array[0][10] = 0 # 如果数组为空，存入0
else:
    empty_2d_array[0][10] = f"{non_zero_D.values[0]}-{D_values[0]} {non_zero_D.values[D_values[1]]}-{non_zero_D.values[-1]}"  # 如果不按升序排列，存入分段的交易时段

non_zero_L = df['服务费(元)'][df['服务费(元)'] != 0]  # 获取非零的服务费
empty_2d_array[0][13] = non_zero_L.sum()  # 计算总服务费并存入数组
empty_2d_array[0][12] = 0 if non_zero_I.sum() == 0 else float(Decimal(non_zero_L.sum()) / Decimal(non_zero_I.sum()))  # 计算单位服务费并存入数组
empty_2d_array[0][14] = 1 if D_values is None else 0 if D_values == 0 else 2  # 判断交易时段是否按升序排列并存入数组

# 定义Excel工作簿路径
excel_workbook_path = "/Users/hang/Downloads/湖南邦锦能源科技有限公司/储存/2025福冲储能站深度调峰统计分析表.xlsx"
workbook = openpyxl.load_workbook(excel_workbook_path)  # 打开工作簿
month = empty_2d_array[0][0][5:6] + "月份"  # 获取月份
sheet = next((workbook[name] for name in workbook.sheetnames if month in name), None)  # 获取包含月份的工作表
if sheet is None:
    raise KeyError(f"No worksheet containing {month} found.")  # 如果工作表不存在，抛出异常

the_day = empty_2d_array[0][0][7:8].lstrip("0")  # 获取日期
row_number = next((cell.row for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row) for cell in row if cell.value == int(the_day)), None)  # 获取日期对应的行号
if row_number is None:
    raise ValueError(f"No row found for day {the_day}")  # 如果行号不存在，抛出异常

for col_index, value in enumerate(empty_2d_array[0], start=1):  # 遍历数组
    sheet.cell(row=row_number, column=col_index + 1, value=value)  # 将数组值写入工作表

workbook.save(excel_workbook_path)  # 保存工作簿
print(empty_2d_array)  # 打印数组

