# 操作系统交互模块
import os  # 提供与操作系统交互的功能，如文件路径操作、环境变量访问等
 
# 字符串处理模块
import re  # 提供正则表达式支持，用于字符串匹配、查找、替换等操作
 
# 时间处理模块
import time  # 提供时间相关的函数，如获取当前时间、时间戳转换等
from datetime import datetime, timedelta  # 提供日期和时间处理功能，包括日期时间对象和时间差计算
 
# 子进程管理模块
import subprocess  # 用于执行外部命令或程序，并获取其输出
 
# Excel文件处理模块
from openpyxl import load_workbook  # 用于加载和读取Excel文件（.xlsx格式）

# 获取 TextEdit 中的内容
def get_textedit_content():
    script = '''
    tell application "TextEdit"
        activate
        set theText to the text of the front document
    end tell
    return theText
    '''
    result = subprocess.run(['osascript', '-e', script], capture_output=True, text=True)
    return result.stdout.strip()

# 关闭 TextEdit
def close_textedit():
    script = '''
    tell application "TextEdit"
        quit saving no
    end tell
    '''
    subprocess.run(['osascript', '-e', script])

# 定义函数，查找符合日期模式的文件
def find_date_files(folder_path, pattern_str):
    pattern = re.compile(pattern_str)  # 编译正则表达式模式
    return [os.path.join(root, filename) for root, _, files in os.walk(folder_path) for filename in files if pattern.match(filename)]  # 返回符合模式的文件路径列表

def main_function():
    # 要打开的.doc文件路径
    file_path = '/Users/hang/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/237e6ba0b2b7af52a974319f16a23be7/Message/MessageTemp/c1c30910da9401121316d3c0d92ec2ed/File/'
    word_str = r'^\d{4}-\d{2}-\d{2}RB\.doc$'
    word_doc_path = find_date_files(file_path, word_str)  # 查找符合模式的Word文件路径
    # 获取今天的日期并格式化为文件名匹配的模式
    yesterday = (datetime.today() - timedelta(days=1)).strftime('%Y-%m-%d')
    yesterday_file_name = f"{yesterday}RB.doc"

    # 检查找到的文件名是否与前一天的日期匹配
    desired_file_path = None
    for path in word_doc_path:
        file_name = os.path.basename(path)
        if file_name == yesterday_file_name:
            desired_file_path = path
            break
    
    subprocess.run(['open', '-a', 'TextEdit', desired_file_path]) # 打开文件

    # 休眠几秒
    time.sleep(0.3)
    # 获取 TextEdit 中的内容
    txt_content = get_textedit_content()
    # 关闭 TextEdit
    close_textedit()

    # 正则表达式匹配模式
    pattern_datetime = re.compile(r"\d{4}年\d{2}月\d{2}日")
    pattern_windpower = re.compile(r"风电分摊(\d+(\.\d+)?)万元（占比(\d+(\.\d+)?)%，度电分摊费用为(\d+(\.\d+)?)元/千瓦时），")
    pattern_photovoltaic = re.compile(r"省调光伏总分摊费(\d+(\.\d+)?)万元（占比(\d+(\.\d+)?)%，度电分摊费用为(\d+(\.\d+)?)元/千瓦时）。")

    # 创建空列表
    data = [None] * 8

    # 提取日期
    date_match = pattern_datetime.search(txt_content)
    if date_match:
        date_obj = datetime.strptime(date_match.group(), "%Y年%m月%d日")
        data[0] = date_obj.strftime("%Y/%m/%d").replace("/0", "/").replace("/0", "/")
        weekdays = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
        data[1] = weekdays[date_obj.weekday()]

    # 提取风电数据
    windpower_match = pattern_windpower.search(txt_content)
    if windpower_match:
        data[2] = windpower_match.group(1)
        data[3] = windpower_match.group(3)
        data[4] = windpower_match.group(5)

    # 提取光伏数据
    photovoltaic_match = pattern_photovoltaic.search(txt_content)
    if photovoltaic_match:
        data[5] = photovoltaic_match.group(1)
        data[6] = photovoltaic_match.group(3)
        data[7] = photovoltaic_match.group(5)

    print(data)

    # 读取Excel文件
    excel_workbook_path = "/Users/hang/Downloads/湖南邦锦能源科技有限公司/储存/2025年湖南省风电光伏辅助服务市场统计.xlsx"
    workbook = load_workbook(excel_workbook_path)
        
    # 查找对应月份的工作表
    month = data[0].split('/')[1] + "月"
    sheet_name = next((name for name in workbook.sheetnames if month in name), None)
    if sheet_name is None:
        raise KeyError(f"No worksheet containing {month} found.")
        
    sheet = workbook[sheet_name]
    the_day = int(data[0].split('/')[2])
        
    # 查找对应日期的行
    row_index = None
    for rowx in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if rowx[0] == the_day:
            row_index = rowx[0]+2
            break
    if row_index is None:
        raise ValueError(f"Day {the_day} not found in the first column of the sheet.")
        
    # 更新Excel表格中的数据
    sheet.cell(row=row_index, column=2, value=data[0])
    sheet.cell(row=row_index, column=3, value=data[1])
    sheet.cell(row=row_index, column=4, value=float(data[3]) / 100 if data[3] is not None else None)
    sheet.cell(row=row_index, column=5, value=float(data[2]) if data[2] is not None else None)
    sheet.cell(row=row_index, column=6, value=float(data[4]) if data[4] is not None else None)
    sheet.cell(row=row_index, column=7, value=float(data[6]) / 100 if data[6] is not None else None)
    sheet.cell(row=row_index, column=8, value=float(data[5]) if data[5] is not None else None)
    sheet.cell(row=row_index, column=9, value=float(data[7]) if data[7] is not None else None)
        
    # 保存更新后的Excel文件
    workbook.save(excel_workbook_path)

if __name__ == "__main__":
    main_function()