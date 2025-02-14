import subprocess  # 用于执行外部命令
 
# Excel文件读取模块
import xlrd  # 用于读取旧版本的Excel文件（.xls）
import openpyxl  # 用于读取和写入新版本的Excel文件（.xlsx）
 
import time  # 用于处理时间相关的功能
 
# 字符串处理模块
from difflib import get_close_matches  # 用于查找字符串列表中与目标字符串最接近的匹配项
import re  # 用于处理正则表达式，匹配、查找、替换字符串等
 
import os  # 提供与操作系统交互的功能，如文件路径操作、环境变量访问等
 
# 数据处理模块
import pandas as pd  # 提供高性能、易用的数据结构和数据分析工具
 
# 文档处理模块
from docx import Document  # 用于创建、修改Word文档
 
# 日期时间处理模块
from datetime import datetime, timedelta  # datetime用于处理日期和时间，timedelta用于计算时间差
 
# 精确数值处理模块
from decimal import Decimal  # 用于执行精确的浮点数运算，避免浮点数精度问题

# 定义函数，查找符合日期模式的文件
def find_date_files(folder_path, pattern_str):
    pattern = re.compile(pattern_str)  # 编译正则表达式模式
    return [os.path.join(root, filename) for root, _, files in os.walk(folder_path) for filename in files if pattern.match(filename)]  # 返回符合模式的文件路径列表

# 定义函数，判断数组是否按升序排列
def ascending_order_judgment(lst):
    if len(lst) == 0:  # 如果列表为空，返回0
        return 0
    for i in range(1, len(lst)):  # 遍历列表
        if lst[i] != lst[i - 1] + 1:  # 如果当前元素不等于前一个元素加1
            return [lst[i - 1], i]  # 返回前一个元素和当前索引
    return None  # 如果列表按升序排列，返回None

# 获取 TextEdit 中的内容
def get_textedit_content():
    script = '''
    tell application "TextEdit"
        activate
        set theText to the text of the front document
    end tell
    return theText
    '''
    result = subprocess.run(['osascript', '-e', script], capture_output=True, text=True)
    return result.stdout.strip()

# 关闭 TextEdit
def close_textedit():
    script = '''
    tell application "TextEdit"
        quit saving no
    end tell
    '''
    subprocess.run(['osascript', '-e', script])


# 定义Word文档路径和文件名模式
file_path = "/Users/hang/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/237e6ba0b2b7af52a974319f16a23be7/Message/MessageTemp/c1c30910da9401121316d3c0d92ec2ed/File/"
word_str = r'^\d{4}-\d{2}-\d{2}RB\.doc$'
word_doc_path = find_date_files(file_path, word_str)  # 查找符合模式的Word文件路径

# 获取今天的日期并格式化为文件名匹配的模式
yesterday = (datetime.today() - timedelta(days=1)).strftime('%Y-%m-%d')
yesterday_file_name = f"{yesterday}RB.doc"

# 检查找到的文件名是否与前一天的日期匹配
desired_file_path = None
for path in word_doc_path:
    file_name = os.path.basename(path)
    if file_name == yesterday_file_name:
        desired_file_path = path
        break

# 打开 Word 文档
subprocess.run(['open', '-a', 'TextEdit', desired_file_path])
# 休眠几秒
time.sleep(0.3)
# 获取 TextEdit 中的内容
txt_content = get_textedit_content()
# 关闭 TextEdit
close_textedit()# 获取Word文档的内容

# 定义正则表达式模式，用于匹配日期、总服务费和储能电站服务费
pattern_datetime = re.compile(r"\d{4}年\d{2}月\d{2}日")
pattern_total_amount = re.compile(r"\d{4}年\d{2}月\d{2}日调峰辅助服务市场产生总服务费(\d+(\.\d+)?)万元。")
pattern_energy_storage = re.compile(r"，储能电站总服务费为(\d+(\.\d+)?)万元（占比(\d+(\.\d+)?)%）。")

empty_2d_array = [None] * 17  # 创建一个空的一维数组

date_match = pattern_datetime.search(txt_content)  # 在Word内容中查找日期
if date_match:
    date_obj = datetime.strptime(date_match.group(), "%Y年%m月%d日")  # 将匹配到的日期字符串转换为日期对象
    weekday_name = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"][date_obj.weekday()]  # 获取星期名称
    empty_2d_array[1] = weekday_name  # 将星期名称存入数组
    empty_2d_array[0] = date_obj.strftime("%Y/%m/%d").replace("/0", "/")  # 将日期格式化为字符串并存入数组

total_amount_match = pattern_total_amount.search(txt_content)  # 在Word内容中查找总服务费
if total_amount_match:
    empty_2d_array[5] = float(total_amount_match.group(1))  # 将匹配到的总服务费存入数组

energy_storage_match = pattern_energy_storage.search(txt_content)  # 在Word内容中查找储能电站服务费占比
if energy_storage_match:
    empty_2d_array[6] = float(Decimal(energy_storage_match.group(3)) / Decimal(100))  # 将匹配到的储能电站服务费占比存入数组

# 定义Excel文档路径和文件名模式
excel_doc_path = "/Users/hang/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/237e6ba0b2b7af52a974319f16a23be7/Message/MessageTemp/c1c30910da9401121316d3c0d92ec2ed/File/"
excel_report_str = r'肖家湾储能电站运行日报.*.xlsx'
excel_settlement_str = r'深度调峰周期结算数据.*(\d{1,2}月\d{1,2}日).xls'
matched_report_files = find_date_files(excel_doc_path, excel_report_str)  # 查找符合模式的运行日报文件路径
matched_settlement_str = find_date_files(excel_doc_path, excel_settlement_str)  # 查找符合模式的结算数据文件路径

formatted_date_simple = date_match.group()[date_match.group().find("年") + 1:]
sheet_name = formatted_date_simple  # 将处理后的日期字符串作为工作表名称

if not matched_report_files:
    raise FileNotFoundError("No matching report files found.")
workbook = openpyxl.load_workbook(matched_report_files[0])  # 打开第一个符合条件的运行日报文件
closest_matches = get_close_matches(sheet_name, workbook.sheetnames, n=1, cutoff=0.6)
if not closest_matches:
    raise ValueError(f"No close match found for sheet '{sheet_name}' in the workbook.")  # 如果没有找到接近的工作表，抛出异常
sheet = workbook[closest_matches[0]]  # 获取最接近的工作表

empty_2d_array[2] = sheet['G4'].value  # 获取单元格F4的值并存入数组
empty_2d_array[3] = int(sheet['K4'].value.replace("℃", ""))  # 获取单元格H4的值并去掉所有的℃后存入数组
empty_2d_array[4] = sheet['N4'].value  # 获取单元格J4的值并存入数组

empty_2d_array[7] = float(Decimal(sheet['A6'].value))  # 获取单元格A6的值并转换为浮点数后存入数组
empty_2d_array[8] = float(Decimal(sheet['B6'].value))  # 获取单元格B6的值并转换为浮点数后存入数组

# 打开第一个符合条件的结算数据文件
workbook = xlrd.open_workbook(matched_settlement_str[0])
sheet = workbook.sheet_by_index(0)  # 获取第一个工作表

# 获取表头行
header = sheet.row_values(2)

# 获取列索引
index_I = header.index('计算深调电量(MWH)')
index_D = header.index('交易时段')
index_J = header.index('服务费调节系数')
index_L = header.index('服务费(元)')

# 获取非零的深调电量
# 获取非零的深调电量和交易时段
non_zero_I = []  # 创建一个空列表，用于存储非零的深调电量
non_zero_D = []  # 创建一个空列表，用于存储非零的交易时段
non_zero_L = []  # 创建一个空列表，用于存储非零的服务费
for row in range(3, sheet.nrows):  # 遍历工作表的每一行，从第4行开始
    cell_value_I = sheet.cell_value(row, index_I)  # 获取当前行的深调电量
    if cell_value_I != '0':  # 如果深调电量不为零
        non_zero_I.append(cell_value_I)  # 将深调电量转换为浮点数并添加到列表中
        non_zero_D.append(int(sheet.cell_value(row, index_D)))  # 将交易时段添加到列表中
        non_zero_L.append(sheet.cell_value(row, index_L))  # 将服务费添加到列表中

# 计算总深调电量并存入数组
empty_2d_array[9] = sum(map(lambda x: float(Decimal(x) * Decimal('1000')), non_zero_I))

# 将交易时段存入数组
D_values = ascending_order_judgment(non_zero_D)  # 判断交易时段是否按升序排列
if D_values is None:
    empty_2d_array[10] = f"{non_zero_D[0]}-{non_zero_D[-1]}"  # 如果按升序排列，存入数组
elif D_values == 0:
    empty_2d_array[10] = 0  # 如果数组为空，存入0
else:
    empty_2d_array[10] = f"{non_zero_D[0]}-{D_values[0]} {non_zero_D[D_values[1]]}-{non_zero_D[-1]}"  # 如果不按升序排列，存入分段的交易时段

# 计算总服务费并存入数组
empty_2d_array[14] = sum(map(lambda x: float(Decimal(x)), non_zero_L))
# 将服务费调节系数存入数组
empty_2d_array[12] = float(sheet.cell_value(4,index_J))
# 计算中标单价存入数组
empty_2d_array[13] = 0 if empty_2d_array[14] == 0 else empty_2d_array[14] / sum(map(lambda x: float(Decimal(x)), non_zero_I))
# 计算总盘占比并存入数组
empty_2d_array[15] = 0 if empty_2d_array[6] == 0 else empty_2d_array[14] / (empty_2d_array[5] * 10000 * empty_2d_array[6])
# 判断调用次数并存入数组
empty_2d_array[16] = 1 if D_values is None else 0 if D_values == 0 else 2  

# 定义Excel工作簿路径
excel_workbook_path = "/Users/hang/Downloads/湖南邦锦能源科技有限公司/储存/2025肖家湾储能站深度调峰统计分析表.xlsx"
workbook = openpyxl.load_workbook(excel_workbook_path)  # 打开工作簿
month = empty_2d_array[0][5:6].lstrip("0") + "月份"  # 获取月份
sheet = next((workbook[name] for name in workbook.sheetnames if month in name), None)  # 获取包含月份的工作表
if sheet is None:
    raise KeyError(f"No worksheet containing {month} found.")  # 如果工作表不存在，抛出异常

the_day = empty_2d_array[0][7:10].lstrip("0")  # 获取日期
row_number = next((cell.row for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row) for cell in row if cell.value == int(the_day)), None)  # 获取日期对应的行号
if row_number is None:
    raise ValueError(f"No row found for day {the_day}")  # 如果行号不存在，抛出异常

for col_index, value in enumerate(empty_2d_array, start=1):  # 遍历数组
    if value is not None:  # 检查值是否为None
        sheet.cell(row=row_number, column=col_index + 1, value=value)  # 将数组值写入工作表

workbook.save(excel_workbook_path)  # 保存工作簿

print(empty_2d_array)  # 打印数组
