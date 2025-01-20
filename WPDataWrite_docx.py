import re
import os
import pandas as pd
from docx import Document
from datetime import datetime
from openpyxl import load_workbook

# 查找符合日期格式的RB.docx文件
def find_date_rb_docx_files(folder_path):
    pattern = re.compile(r'^\d{4}-\d{2}-\d{2}RB\.docx$')
    matching_files = []
    for root, dirs, files in os.walk(folder_path):
        for filename in files:
            if pattern.match(filename):
                matching_files.append(os.path.join(root, filename))
    return matching_files

def main_function():
    # Word文档路径
    word_doc_path = "/Users/hang/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/237e6ba0b2b7af52a974319f16a23be7/Message/MessageTemp/c1c30910da9401121316d3c0d92ec2ed/File/"
    if not os.path.exists(word_doc_path):
        raise FileNotFoundError(f"The directory {word_doc_path} does not exist.")
    word_docx_path = find_date_rb_docx_files(word_doc_path)

    if not word_docx_path:
        raise FileNotFoundError("No matching Word documents found in the specified directory.")

    # 读取Word文档内容
    doc = Document(word_docx_path[0])
    full_text = [para.text for para in doc.paragraphs]
    word_content = '\n'.join(full_text)

    # 正则表达式匹配模式
    pattern_datetime = re.compile(r"\d{4}年\d{2}月\d{2}日")
    pattern_windpower = re.compile(r"风电分摊(\d+(\.\d+)?)万元（占比(\d+(\.\d+)?)%，度电分摊费用为(\d+(\.\d+)?)元/千瓦时），")
    pattern_photovoltaic = re.compile(r"省调光伏总分摊费(\d+(\.\d+)?)万元（占比(\d+(\.\d+)?)%，度电分摊费用为(\d+(\.\d+)?)元/千瓦时）。")

    data = [None] * 8

    # 提取日期
    date_match = pattern_datetime.search(word_content)
    if date_match:
        date_obj = datetime.strptime(date_match.group(), "%Y年%m月%d日")
        data[0] = date_obj.strftime("%Y/%m/%d").replace("/0", "/").replace("/0", "/")
        weekdays = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
        data[1] = weekdays[date_obj.weekday()]

    # 提取风电数据
    windpower_match = pattern_windpower.search(word_content)
    if windpower_match:
        data[2] = windpower_match.group(1)
        data[3] = windpower_match.group(3)
        data[4] = windpower_match.group(5)

    # 提取光伏数据
    photovoltaic_match = pattern_photovoltaic.search(word_content)
    if photovoltaic_match:
        data[5] = photovoltaic_match.group(1)
        data[6] = photovoltaic_match.group(3)
        data[7] = photovoltaic_match.group(5)

    print(data)

    # 读取Excel文件
    excel_workbook_path = "/Users/hang/Downloads/湖南邦锦能源科技有限公司/储存/2025年湖南省风电光伏辅助服务市场统计.xlsx"
    workbook = load_workbook(excel_workbook_path)
    
    # 查找对应月份的工作表
    month = data[0].split('/')[1] + "月"
    sheet_name = next((name for name in workbook.sheetnames if month in name), None)
    if sheet_name is None:
        raise KeyError(f"No worksheet containing {month} found.")
    
    sheet = workbook[sheet_name]
    the_day = int(data[0].split('/')[2])
    
    # 查找对应日期的行
    row_index = None
    for rowx in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if rowx[0] == the_day:
            row_index = rowx[0]+2
            break
    if row_index is None:
        raise ValueError(f"Day {the_day} not found in the first column of the sheet.")
    
    # 更新Excel表格中的数据
    sheet.cell(row=row_index, column=2, value=data[0])
    sheet.cell(row=row_index, column=3, value=data[1])
    sheet.cell(row=row_index, column=4, value=float(data[3]) / 100 if data[3] is not None else None)
    sheet.cell(row=row_index, column=5, value=float(data[2]) if data[2] is not None else None)
    sheet.cell(row=row_index, column=6, value=float(data[4]) if data[4] is not None else None)
    sheet.cell(row=row_index, column=7, value=float(data[6]) / 100 if data[6] is not None else None)
    sheet.cell(row=row_index, column=8, value=float(data[5]) if data[5] is not None else None)
    sheet.cell(row=row_index, column=9, value=float(data[7]) if data[7] is not None else None)
    
    # 保存更新后的Excel文件
    workbook.save(excel_workbook_path)
if __name__ == "__main__":
    main_function()
